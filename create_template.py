# create_template.py
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Cria workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Serviços Públicos"

# Cabeçalhos
headers = [
    "nome",
    "tipo",
    "latitude",
    "longitude",
    "endereco",
    "telefone",
    "horario_funcionamento",
    "descricao"
]

# Estilização do cabeçalho
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)

# Adiciona cabeçalhos com estilo
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Ajusta largura das colunas
ws.column_dimensions['A'].width = 40  # nome
ws.column_dimensions['B'].width = 15  # tipo
ws.column_dimensions['C'].width = 12  # latitude
ws.column_dimensions['D'].width = 12  # longitude
ws.column_dimensions['E'].width = 50  # endereco
ws.column_dimensions['F'].width = 18  # telefone
ws.column_dimensions['G'].width = 25  # horario
ws.column_dimensions['H'].width = 50  # descricao

# Adiciona exemplos
exemplos = [
    [
        "Hospital da Restauração",
        "Saúde",
        -8.0524,
        -34.8813,
        "Av. Gov. Agamenon Magalhães, s/n - Derby",
        "(81) 3184-1300",
        "24 horas",
        "Hospital de referência em trauma e urgências"
    ],
    [
        "Detran Recife",
        "Transporte",
        -8.0476,
        -34.8770,
        "Av. Agamenon Magalhães, 2200 - Boa Vista",
        "(81) 3184-9000",
        "Segunda a Sexta: 8h às 17h",
        "Atendimento para habilitação, veículos e infrações"
    ],
    [
        "Faculdade Senac Pernambuco",
        "Educação",
        -8.0584,
        -34.8811,
        "R. do Pombal, 57 - Santo Amaro",
        "(81) 3413-6666",
        "Segunda a Sexta: 8h às 22h",
        "Instituição de ensino superior e técnico"
    ]
]

# Estilo para exemplos (fundo amarelo claro)
exemplo_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

for row_num, exemplo in enumerate(exemplos, 2):
    for col_num, valor in enumerate(exemplo, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = valor
        cell.fill = exemplo_fill

# Adiciona 20 linhas vazias para preencher
for row_num in range(5, 26):
    for col_num in range(1, 9):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = ""

# Adiciona instruções em uma aba separada
ws_instrucoes = wb.create_sheet("Instruções")
ws_instrucoes.column_dimensions['A'].width = 100

instrucoes = [
    ["📋 INSTRUÇÕES PARA PREENCHER A PLANILHA"],
    [""],
    ["1. CAMPOS OBRIGATÓRIOS (não podem estar vazios):"],
    ["   • nome: Nome completo do serviço público"],
    ["   • tipo: Categoria do serviço (Saúde, Transporte, Educação, Segurança, etc)"],
    ["   • latitude: Coordenada de latitude (ex: -8.0524)"],
    ["   • longitude: Coordenada de longitude (ex: -34.8813)"],
    [""],
    ["2. CAMPOS OPCIONAIS (podem ficar vazios):"],
    ["   • endereco: Endereço completo do serviço"],
    ["   • telefone: Telefone de contato"],
    ["   • horario_funcionamento: Horário de atendimento"],
    ["   • descricao: Descrição breve do serviço"],
    [""],
    ["3. TIPOS DE SERVIÇOS SUGERIDOS:"],
    ["   • Saúde (hospitais, UPAs, postos de saúde)"],
    ["   • Transporte (Detran, terminais, estações de metrô)"],
    ["   • Educação (escolas, universidades, faculdades)"],
    ["   • Segurança (delegacias, corpo de bombeiros)"],
    ["   • Assistência Social (CRAS, CREAS)"],
    ["   • Administração (prefeitura, fóruns, tribunais)"],
    ["   • Outros (INSS, Correios, Receita Federal)"],
    [""],
    ["4. COMO OBTER COORDENADAS (latitude/longitude):"],
    ["   • Abra o Google Maps: https://www.google.com/maps"],
    ["   • Procure o endereço do serviço"],
    ["   • Clique com botão direito no local exato"],
    ["   • Clique nas coordenadas que aparecem (ex: -8.0524, -34.8813)"],
    ["   • As coordenadas serão copiadas automaticamente"],
    ["   • Cole na planilha (latitude na coluna C, longitude na coluna D)"],
    [""],
    ["5. APÓS PREENCHER:"],
    ["   • Salve o arquivo Excel"],
    ["   • Acesse: http://localhost:8000/docs"],
    ["   • Procure: POST /servicos/importar-csv"],
    ["   • Clique em 'Try it out'"],
    ["   • Clique em 'Choose File' e selecione seu arquivo"],
    ["   • Clique em 'Execute'"],
    ["   • Aguarde a importação finalizar"],
    [""],
    ["6. EXEMPLOS:"],
    ["   Veja a aba 'Serviços Públicos' - linhas 2, 3 e 4"],
    ["   (linhas com fundo amarelo são exemplos)"],
    ["   Você pode deletar os exemplos ou mantê-los"],
    [""],
    ["✅ IMPORTANTE: NÃO altere os nomes dos cabeçalhos da linha 1!"],
    ["✅ DICA: Use o Google Maps para encontrar as coordenadas exatas"]
]

for row_num, linha in enumerate(instrucoes, 1):
    cell = ws_instrucoes.cell(row=row_num, column=1)
    cell.value = linha[0]
    if row_num == 1:
        cell.font = Font(bold=True, size=14, color="0563C1")
    elif "OBRIGATÓRIOS" in str(linha[0]) or "OPCIONAIS" in str(linha[0]) or "TIPOS" in str(linha[0]) or "COMO OBTER" in str(linha[0]) or "APÓS" in str(linha[0]) or "EXEMPLOS" in str(linha[0]):
        cell.font = Font(bold=True, size=11)

# Salva arquivo
wb.save("template_servicos_publicos_recife.xlsx")
print("✅ Arquivo criado com sucesso!")
print("📁 Nome: template_servicos_publicos_recife.xlsx")
print("📍 Localização: pasta atual do projeto")
print("\n🎯 Próximos passos:")
print("1. Abra o arquivo Excel")
print("2. Preencha com os serviços públicos")
print("3. Salve o arquivo")
print("4. Faça upload pelo Swagger em: POST /servicos/importar-csv")